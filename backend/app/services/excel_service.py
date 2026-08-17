"""Excel (.xlsx) тайлан үүсгэх үйлчилгээ — openpyxl (WP8).

Бүх толгой мөр монгол хэлээр, тод үсэг + дэвсгэр өнгөтэй, мөр царцаасан
(``freeze_panes``), баганын өргөн тохируулсан. Мөнгө ``#,##0.00``, литр
``#,##0.000`` форматтай.

Функц бүр ``bytes`` буцаана — router нь ``xlsx_response`` -оор дамжуулж
файл болгон илгээнэ.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any
from urllib.parse import quote

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MONEY_FMT = "#,##0.00"
LITER_FMT = "#,##0.000"
INT_FMT = "#,##0"
PCT_FMT = "#,##0.00"
DATE_FMT = "YYYY-MM-DD"
DATETIME_FMT = "YYYY-MM-DD HH:MM"

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="2563EB", end_color="2563EB")
TITLE_FONT = Font(bold=True, size=13)
TOTAL_FONT = Font(bold=True)
TOTAL_FILL = PatternFill("solid", start_color="E2E8F0", end_color="E2E8F0")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="CBD5E1")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --------------------------------------------------------------------------- #
# Дотоод туслахууд
# --------------------------------------------------------------------------- #
def _get(source: Any, key: str, default: Any = None) -> Any:
    """dict эсвэл объектоос талбар унших нэгдсэн хандалт."""
    if source is None:
        return default
    if isinstance(source, Mapping):
        return source.get(key, default)
    return getattr(source, key, default)


def _cell_value(value: Any) -> Any:
    """openpyxl-д тааруулсан утга (Decimal, огноо шууд, бусад нь текст)."""
    if value is None:
        return None
    if isinstance(value, bool):
        return "Тийм" if value else "Үгүй"
    if isinstance(value, (int, float, Decimal)):
        return value
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo is not None else value
    if isinstance(value, date):
        return value
    return str(value)


def _apply_widths(ws: Worksheet, widths: Sequence[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _write_header(ws: Worksheet, headers: Sequence[str], row: int = 1) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    ws.row_dimensions[row].height = 26


def _write_row(
    ws: Worksheet,
    row: int,
    values: Sequence[Any],
    formats: Sequence[str | None],
    *,
    bold: bool = False,
    fill: PatternFill | None = None,
) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=_cell_value(value))
        fmt = formats[col - 1] if col - 1 < len(formats) else None
        if fmt:
            cell.number_format = fmt
        if bold:
            cell.font = TOTAL_FONT
        if fill is not None:
            cell.fill = fill
        cell.border = CELL_BORDER


def _build_sheet(
    ws: Worksheet,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    formats: Sequence[str | None],
    widths: Sequence[int],
    *,
    total_row: Sequence[Any] | None = None,
) -> None:
    _write_header(ws, headers)
    line = 2
    for values in rows:
        _write_row(ws, line, values, formats)
        line += 1
    last_data_row = line - 1
    if total_row is not None:
        _write_row(ws, line, total_row, formats, bold=True, fill=TOTAL_FILL)
    _apply_widths(ws, widths)
    ws.freeze_panes = "A2"
    # Шүүлтэд нийт дүнгийн мөр орохгүй.
    last_column = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_column}{max(last_data_row, 1)}"


def _new_workbook(first_sheet_title: str) -> tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = wb.active
    ws.title = first_sheet_title
    return wb, ws


def _to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _period_text(data: Any) -> str:
    date_from = _get(data, "date_from")
    date_to = _get(data, "date_to")
    if date_from and date_to:
        return f"{date_from} — {date_to}"
    if date_from:
        return str(date_from)
    return ""


def _strip_tz(value: Any) -> Any:
    """Excel timezone-той огноо дэмждэггүй тул бүсийн мэдээллийг хасна."""
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


# --------------------------------------------------------------------------- #
# Борлуулалтын тайлан
# --------------------------------------------------------------------------- #
def sales_report_xlsx(data: Any) -> bytes:
    """``report_service.sales_summary`` -ийн үр дүнг Excel болгоно."""
    wb, ws = _new_workbook("Борлуулалт")

    headers = [
        "Хугацаа",
        "Гүйлгээний тоо",
        "Нийт дүн",
        "НӨАТ",
        "Түлшний дүн",
        "Дэлгүүрийн дүн",
        "Литр",
        "Өртөг",
        "Нийт ашиг",
    ]
    formats: list[str | None] = [
        None,
        INT_FMT,
        MONEY_FMT,
        MONEY_FMT,
        MONEY_FMT,
        MONEY_FMT,
        LITER_FMT,
        MONEY_FMT,
        MONEY_FMT,
    ]
    widths = [16, 16, 16, 14, 16, 16, 14, 16, 16]

    rows = []
    for row in _get(data, "rows", []) or []:
        rows.append(
            [
                _get(row, "period"),
                _get(row, "sale_count", 0),
                _get(row, "total"),
                _get(row, "vat"),
                _get(row, "fuel_total"),
                _get(row, "store_total"),
                _get(row, "liters"),
                _get(row, "cogs"),
                _get(row, "gross_profit"),
            ]
        )

    totals = _get(data, "totals") or {}
    total_row = [
        "Нийт",
        _get(totals, "sale_count", 0),
        _get(totals, "total"),
        _get(totals, "vat"),
        _get(totals, "fuel_total"),
        _get(totals, "store_total"),
        _get(totals, "liters"),
        _get(totals, "cogs"),
        _get(totals, "gross_profit"),
    ]

    _build_sheet(ws, headers, rows, formats, widths, total_row=total_row)

    meta = wb.create_sheet("Тайлбар")
    meta["A1"] = "Борлуулалтын тайлан"
    meta["A1"].font = TITLE_FONT
    meta["A3"] = "Хугацаа"
    meta["B3"] = _period_text(data)
    meta["A4"] = "Бүлэглэлт"
    meta["B4"] = _get(data, "granularity_name") or _get(data, "granularity") or ""
    meta["A5"] = "Дундаж чек"
    meta["B5"] = _get(totals, "avg_check")
    meta["B5"].number_format = MONEY_FMT
    meta["A6"] = "Ашгийн хувь"
    meta["B6"] = _get(totals, "margin_pct")
    meta["B6"].number_format = PCT_FMT
    _apply_widths(meta, [22, 28])

    return _to_bytes(wb)


# --------------------------------------------------------------------------- #
# Түлшний тайлан
# --------------------------------------------------------------------------- #
def fuel_report_xlsx(data: Any) -> bytes:
    """``report_service.fuel_report`` -ийн үр дүнг Excel болгоно."""
    wb, ws = _new_workbook("Түлшний төрөл")

    headers = ["Код", "Нэр", "Литр", "Орлого", "Орлого (НӨАТ-гүй)", "Өртөг", "Ашиг", "Ашгийн %"]
    formats: list[str | None] = [
        None,
        None,
        LITER_FMT,
        MONEY_FMT,
        MONEY_FMT,
        MONEY_FMT,
        MONEY_FMT,
        PCT_FMT,
    ]
    widths = [12, 24, 14, 16, 20, 16, 16, 12]

    rows = []
    for row in _get(data, "grades", []) or []:
        rows.append(
            [
                _get(row, "code"),
                _get(row, "name"),
                _get(row, "liters"),
                _get(row, "revenue"),
                _get(row, "revenue_net"),
                _get(row, "cogs"),
                _get(row, "margin"),
                _get(row, "margin_pct"),
            ]
        )

    totals = _get(data, "grade_totals") or {}
    total_row = [
        "Нийт",
        "",
        _get(totals, "liters"),
        _get(totals, "revenue"),
        _get(totals, "revenue_net"),
        _get(totals, "cogs"),
        _get(totals, "margin"),
        _get(totals, "margin_pct"),
    ]
    _build_sheet(ws, headers, rows, formats, widths, total_row=total_row)

    pumps_ws = wb.create_sheet("Түгээгүүр")
    pump_headers = ["Түгээгүүр", "Нэр", "Хошуу", "Литр", "Дүн"]
    pump_formats: list[str | None] = [INT_FMT, None, INT_FMT, LITER_FMT, MONEY_FMT]
    pump_rows = []
    for row in _get(data, "pumps", []) or []:
        pump_rows.append(
            [
                _get(row, "pump_number"),
                _get(row, "pump_name"),
                _get(row, "nozzle_number"),
                _get(row, "liters"),
                _get(row, "amount"),
            ]
        )
    pump_totals = _get(data, "pump_totals") or {}
    _build_sheet(
        pumps_ws,
        pump_headers,
        pump_rows,
        pump_formats,
        [12, 24, 12, 14, 16],
        total_row=["Нийт", "", "", _get(pump_totals, "liters"), _get(pump_totals, "amount")],
    )

    return _to_bytes(wb)


# --------------------------------------------------------------------------- #
# Нөөцийн тайлан
# --------------------------------------------------------------------------- #
def inventory_xlsx(data: Any) -> bytes:
    """``statement_service.inventory_valuation`` -ийн үр дүнг Excel болгоно."""
    wb, ws = _new_workbook("Сав")

    tank_headers = ["Сав", "Түлш", "Үлдэгдэл (л)", "Дундаж өртөг", "Нийт дүн"]
    tank_formats: list[str | None] = [None, None, LITER_FMT, MONEY_FMT, MONEY_FMT]
    tank_rows = []
    for row in _get(data, "tanks", []) or []:
        tank_rows.append(
            [
                _get(row, "tank_name") or _get(row, "name"),
                _get(row, "fuel_name_mn") or _get(row, "fuel_name"),
                _get(row, "qty", _get(row, "current_l")),
                _get(row, "avg_cost"),
                _get(row, "value"),
            ]
        )
    _build_sheet(
        ws,
        tank_headers,
        tank_rows,
        tank_formats,
        [24, 20, 16, 18, 18],
        total_row=["Нийт", "", "", "", _get(data, "fuel_value")],
    )

    goods_ws = wb.create_sheet("Бараа")
    goods_headers = ["Код", "Нэр", "Үлдэгдэл", "Дундаж өртөг", "Нийт дүн"]
    goods_formats: list[str | None] = [None, None, LITER_FMT, MONEY_FMT, MONEY_FMT]
    goods_rows = []
    for row in _get(data, "products", []) or []:
        goods_rows.append(
            [
                _get(row, "sku"),
                _get(row, "name_mn") or _get(row, "name"),
                _get(row, "qty", _get(row, "stock_qty")),
                _get(row, "avg_cost"),
                _get(row, "value"),
            ]
        )
    _build_sheet(
        goods_ws,
        goods_headers,
        goods_rows,
        goods_formats,
        [16, 32, 16, 18, 18],
        total_row=["Нийт", "", "", "", _get(data, "goods_value")],
    )

    summary = wb.create_sheet("Нэгтгэл")
    summary["A1"] = "Нөөцийн үнэлгээ"
    summary["A1"].font = TITLE_FONT
    pairs = [
        ("Түлшний нөөц", _get(data, "fuel_value")),
        ("Барааны нөөц", _get(data, "goods_value")),
        ("Нийт нөөц", _get(data, "total_value")),
        ("Дэвтрийн үлдэгдэл — түлш (1301)", _get(data, "ledger_fuel")),
        ("Дэвтрийн үлдэгдэл — бараа (1302)", _get(data, "ledger_goods")),
        ("Зөрүү — түлш", _get(data, "fuel_delta")),
        ("Зөрүү — бараа", _get(data, "goods_delta")),
        ("Зөрүү — нийт", _get(data, "total_delta")),
    ]
    for index, (label, value) in enumerate(pairs, start=3):
        summary.cell(row=index, column=1, value=label)
        cell = summary.cell(row=index, column=2, value=_cell_value(value))
        cell.number_format = MONEY_FMT
    _apply_widths(summary, [36, 20])

    return _to_bytes(wb)


# --------------------------------------------------------------------------- #
# Журналын тайлан
# --------------------------------------------------------------------------- #
def journal_xlsx(data: Any) -> bytes:
    """Журналын бичилтүүдийг мөрөөр нь дэлгэсэн Excel."""
    entries = data
    if isinstance(data, Mapping):
        entries = data.get("items") or data.get("entries") or []
    entries = entries or []

    wb, ws = _new_workbook("Журнал")
    headers = [
        "Бичилт №",
        "Огноо",
        "Гүйлгээний утга",
        "Эх сурвалж",
        "Үйл явдал",
        "Мөр",
        "Данс",
        "Тайлбар",
        "Дебит",
        "Кредит",
    ]
    formats: list[str | None] = [
        INT_FMT,
        DATE_FMT,
        None,
        None,
        None,
        INT_FMT,
        None,
        None,
        MONEY_FMT,
        MONEY_FMT,
    ]
    widths = [12, 14, 36, 16, 20, 8, 12, 32, 16, 16]

    rows: list[list[Any]] = []
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")
    for entry in entries:
        lines = _get(entry, "lines") or []
        for line in lines:
            debit = _get(line, "debit") or Decimal("0.00")
            credit = _get(line, "credit") or Decimal("0.00")
            total_debit += Decimal(str(debit))
            total_credit += Decimal(str(credit))
            rows.append(
                [
                    _get(entry, "entry_no"),
                    _strip_tz(_get(entry, "entry_date")),
                    _get(entry, "description"),
                    _get(entry, "source_type"),
                    _get(entry, "event_type"),
                    _get(line, "line_no"),
                    _get(line, "account_code"),
                    _get(line, "memo"),
                    debit,
                    credit,
                ]
            )

    total_row = ["", "", "Нийт", "", "", "", "", "", total_debit, total_credit]
    _build_sheet(ws, headers, rows, formats, widths, total_row=total_row)
    return _to_bytes(wb)


# --------------------------------------------------------------------------- #
# Орлого үр дүнгийн тайлан
# --------------------------------------------------------------------------- #
def pnl_xlsx(data: Any) -> bytes:
    """``statement_service.income_statement`` -ийн үр дүнг Excel болгоно."""
    wb, ws = _new_workbook("Орлого үр дүн")

    headers = ["Хэсэг", "Дансны код", "Дансны нэр", "Дүн"]
    formats: list[str | None] = [None, None, None, MONEY_FMT]
    widths = [22, 14, 40, 18]

    rows: list[list[Any]] = []
    sections = [
        ("Орлого", _get(data, "revenue", []) or []),
        ("Борлуулсан бүтээгдэхүүний өртөг", _get(data, "cogs", []) or []),
        ("Зардал", _get(data, "expense", []) or []),
    ]
    for section, lines in sections:
        for line in lines:
            rows.append([section, _get(line, "code"), _get(line, "name_mn"), _get(line, "amount")])

    _write_header(ws, headers)
    line_no = 2
    for values in rows:
        _write_row(ws, line_no, values, formats)
        line_no += 1

    summary_rows = [
        ("Нийт орлого", _get(data, "total_revenue")),
        ("Нийт өртөг", _get(data, "total_cogs")),
        ("Нийт ашиг", _get(data, "gross_profit")),
        ("Нийт зардал", _get(data, "total_expense")),
        ("Цэвэр ашиг", _get(data, "net_profit")),
    ]
    line_no += 1
    for label, value in summary_rows:
        _write_row(ws, line_no, [label, "", "", value], formats, bold=True, fill=TOTAL_FILL)
        line_no += 1

    _apply_widths(ws, widths)
    ws.freeze_panes = "A2"

    margins = _get(data, "fuel_margins") or []
    if margins:
        margin_ws = wb.create_sheet("Түлшний ашиг")
        _build_sheet(
            margin_ws,
            ["Түлш", "Орлого", "Өртөг", "Ашиг", "Ашгийн %"],
            [
                [
                    _get(row, "fuel_name_mn") or _get(row, "name"),
                    _get(row, "revenue"),
                    _get(row, "cogs"),
                    _get(row, "margin"),
                    _get(row, "margin_pct"),
                ]
                for row in margins
            ],
            [None, MONEY_FMT, MONEY_FMT, MONEY_FMT, PCT_FMT],
            [24, 18, 18, 18, 12],
        )

    meta = wb.create_sheet("Тайлбар")
    meta["A1"] = "Орлого үр дүнгийн тайлан"
    meta["A1"].font = TITLE_FONT
    meta["A3"] = "Хугацаа"
    meta["B3"] = _period_text(data)
    _apply_widths(meta, [22, 28])

    return _to_bytes(wb)



# --------------------------------------------------------------------------- #
# Бараа материалын тайлан /өртгөөр/
# --------------------------------------------------------------------------- #
def inventory_movement_xlsx(data: Any) -> bytes:
    """Шаталсан мөр бүхий хөдөлгөөний тайлан — задаргаа асаасан бол мөн оруулна."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb, ws = _new_workbook("Бараа материал")

    thin = Side(style="thin", color="B0B7C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="EEF1F5")

    ws["A1"] = "Бараа материалын тайлан /өртгөөр/"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = _period_text(data)
    filter_text = _get(data, "filter_text") or ""
    if filter_text:
        ws["A3"] = f"Шүүлтийн нөхцөл: {filter_text}"
        ws["A3"].font = Font(italic=True, size=9)

    top = 5
    ws.cell(row=top, column=1, value="Код")
    ws.cell(row=top, column=2, value="Нэр")
    for title, col in (("Эхний үлдэгдэл", 3), ("Орлого", 5), ("Зарлага", 7), ("Эцсийн үлдэгдэл", 9)):
        ws.cell(row=top, column=col, value=title)
        ws.merge_cells(start_row=top, start_column=col, end_row=top, end_column=col + 1)
        ws.cell(row=top + 1, column=col, value="Тоо")
        ws.cell(row=top + 1, column=col + 1, value="Дүн")
    ws.cell(row=top, column=11, value="Нэгж өртөг")
    ws.merge_cells(start_row=top, start_column=1, end_row=top + 1, end_column=1)
    ws.merge_cells(start_row=top, start_column=2, end_row=top + 1, end_column=2)
    ws.merge_cells(start_row=top, start_column=11, end_row=top + 1, end_column=11)

    for col in range(1, 12):
        for r in (top, top + 1):
            c = ws.cell(row=r, column=col)
            c.font = Font(bold=True, size=10)
            c.fill = head_fill
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")

    row = top + 2
    for item in _get(data, "rows") or []:
        level = int(_get(item, "level") or 0)
        indent = "    " * level
        ws.cell(row=row, column=1, value=_get(item, "code"))
        name_cell = ws.cell(row=row, column=2, value=f"{indent}{_get(item, 'name')}")
        if level == 0:
            name_cell.font = Font(bold=True)
            ws.cell(row=row, column=1).font = Font(bold=True)
        for col, key in (
            (3, "opening_qty"), (4, "opening_value"),
            (5, "in_qty"), (6, "in_value"),
            (7, "out_qty"), (8, "out_value"),
            (9, "closing_qty"), (10, "closing_value"),
            (11, "unit_cost"),
        ):
            ws.cell(row=row, column=col, value=_cell_value(_get(item, key)))
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = border
        row += 1

        for det in _get(item, "details") or []:
            label = f"{_get(det, 'date')} - {_get(det, 'movement_name')}"
            note = _get(det, "note")
            if note:
                label = f"{label} ({note})"
            cell = ws.cell(row=row, column=2, value=f"{indent}    {label}")
            cell.font = Font(italic=True, size=9)
            for col, key in (
                (5, "in_qty"), (6, "in_value"),
                (7, "out_qty"), (8, "out_value"),
                (9, "balance_qty"), (11, "unit_cost"),
            ):
                ws.cell(row=row, column=col, value=_cell_value(_get(det, key)))
            for col in range(1, 12):
                ws.cell(row=row, column=col).border = border
            row += 1

    totals = _get(data, "totals") or {}
    ws.cell(row=row, column=2, value="Нийт дүн").font = Font(bold=True)
    for col, key in (
        (3, "opening_qty"), (4, "opening_value"),
        (5, "in_qty"), (6, "in_value"),
        (7, "out_qty"), (8, "out_value"),
        (9, "closing_qty"), (10, "closing_value"),
    ):
        c = ws.cell(row=row, column=col, value=_cell_value(_get(totals, key)))
        c.font = Font(bold=True)
    for col in range(1, 12):
        ws.cell(row=row, column=col).border = border

    _apply_widths(ws, [14, 46, 12, 16, 12, 16, 12, 16, 12, 16, 14])
    ws.freeze_panes = ws.cell(row=top + 2, column=1)
    return _to_bytes(wb)



# --------------------------------------------------------------------------- #
# Тайлангийн төв
# --------------------------------------------------------------------------- #
def report_center_xlsx(data: Any) -> bytes:
    """Шаталсан бүлэглэлтэй тайлан + задаргаа."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb, ws = _new_workbook("Тайлан")
    thin = Side(style="thin", color="B0B7C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = _get(data, "report_name") or "Тайлан"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = _period_text(data)
    ws["A3"] = f"Шүүлтийн нөхцөл: {_get(data, 'filter_text') or 'Бүгд'}"
    ws["A3"].font = Font(italic=True, size=9)
    labels = _get(data, "group_by_labels") or []
    ws["A4"] = f"Бүлэглэл: {' → '.join(labels)}" if labels else ""
    ws["A4"].font = Font(italic=True, size=9)

    head = 6
    headers = ["Код", "Нэр", "Тоо", "Дүн", "Гүйлгээ"]
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=head, column=col, value=title)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="EEF1F5")
        c.border = border
        c.alignment = Alignment(horizontal="center")

    row = head + 1
    for item in _get(data, "rows") or []:
        level = int(_get(item, "level") or 0)
        indent = "    " * level
        ws.cell(row=row, column=1, value=_get(item, "code"))
        name_cell = ws.cell(row=row, column=2, value=f"{indent}{_get(item, 'name')}")
        if level == 0:
            name_cell.font = Font(bold=True)
        ws.cell(row=row, column=3, value=_cell_value(_get(item, "qty")))
        ws.cell(row=row, column=4, value=_cell_value(_get(item, "amount")))
        ws.cell(row=row, column=5, value=_get(item, "count"))
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
        row += 1

        for det in _get(item, "details") or []:
            label = f"{_get(det, 'date')} · {_get(det, 'tx_type_name')} {_get(det, 'doc_no')} — {_get(det, 'item_name')}"
            person = _get(det, "employee_name")
            if person and person != "—":
                label = f"{label} ({person})"
            cell = ws.cell(row=row, column=2, value=f"{indent}    {label}")
            cell.font = Font(italic=True, size=9)
            ws.cell(row=row, column=3, value=_cell_value(_get(det, "qty")))
            ws.cell(row=row, column=4, value=_cell_value(_get(det, "amount")))
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
            row += 1

    totals = _get(data, "totals") or {}
    ws.cell(row=row, column=2, value="Нийт дүн").font = Font(bold=True)
    ws.cell(row=row, column=3, value=_cell_value(_get(totals, "qty"))).font = Font(bold=True)
    ws.cell(row=row, column=4, value=_cell_value(_get(totals, "amount"))).font = Font(bold=True)
    ws.cell(row=row, column=5, value=_get(totals, "count")).font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border

    _apply_widths(ws, [18, 58, 14, 18, 10])
    ws.freeze_panes = ws.cell(row=head + 1, column=1)
    return _to_bytes(wb)


# --------------------------------------------------------------------------- #
# HTTP хариу
# --------------------------------------------------------------------------- #
def xlsx_response(content: bytes, filename: str) -> StreamingResponse:
    """Excel файлыг татаж авах хариу (кирилл нэрийг RFC 5987-оор дамжуулна)."""
    if not filename.lower().endswith(".xlsx"):
        filename = f"{filename}.xlsx"
    ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "report.xlsx"
    disposition = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(filename)}"
    return StreamingResponse(
        BytesIO(content),
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": disposition,
            "Content-Length": str(len(content)),
        },
    )
