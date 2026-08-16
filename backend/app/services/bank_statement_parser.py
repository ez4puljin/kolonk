"""Банкны хуулгын Excel файлыг задлан унших.

Хаанбанкны хуулгын бүтэц:

    Мөр 0: Хэвлэсэн огноо / хэрэглэгч / интервал зэрэг нийтлэг мэдээлэл
    Мөр 1: Баганы гарчиг — байрлал нь хувирдаг тул динамикаар танина
    Мөр 2..n: Гүйлгээний мөрүүд
    Сүүлийн мөр: "Нийт дүн:" нийлбэр (хасна)

Энэ модуль өгөгдлийн сангаас бүрэн хамааралгүй — цэвэр функцууд тул unit
тестээр шалгахад хялбар.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

#: Тайлбарт эдгээрийн аль нэг байвал банкны шимтгэл гэж үзнэ.
FEE_KEYWORDS: tuple[str, ...] = (
    "хураамж",
    "шимтгэл",
    "commission",
    "fee",
    "үйлчилгээний төлбөр",
)

#: ПОС-ын тооцоо — Хаанбанк тайлбартаа "SETTLEMENT" гэж бичдэг.
SETTLEMENT_RE = re.compile(r"\bSETTLEMENT\b", re.IGNORECASE)

DATE_YMD_RE = re.compile(r"\b(\d{4})[-./](\d{1,2})[-./](\d{1,2})\b")
DATE_DMY_RE = re.compile(r"\b(\d{1,2})[-./](\d{1,2})[-./](\d{4})\b")

#: Statement_MNT_5301234567.xlsx → ("MNT", "5301234567")
FILENAME_RE = re.compile(r"Statement[_-]([A-Za-z]+)[_-](\d+)")

ZERO = Decimal("0.00")


@dataclass
class ParsedTransaction:
    txn_date: datetime | None
    debit: Decimal
    credit: Decimal
    bank_description: str
    bank_counterpart: str
    is_fee: bool


@dataclass
class ParsedStatement:
    account_number: str = ""
    currency: str = "MNT"
    date_from: date | None = None
    date_to: date | None = None
    filename: str = ""
    transactions: list[ParsedTransaction] = field(default_factory=list)


def is_fee_description(desc: str | None) -> bool:
    text = (desc or "").lower()
    return any(word in text for word in FEE_KEYWORDS)


def is_pos_income(desc: str | None) -> bool:
    return bool(desc) and SETTLEMENT_RE.search(desc or "") is not None


def settlement_description(config_text: str | None, bank_desc: str | None) -> str:
    """ПОС гүйлгээний утга = тохиргооны бичвэр + банкны бичвэр.

    Аль нэг нь хоосон бол нөгөөг нь дангаар нь буцаана.
    """
    cfg = (config_text or "").strip()
    bank = (bank_desc or "").strip()
    if cfg and bank:
        return f"{cfg} {bank}"
    return cfg or bank


def _make_date(year: int, month: int, day: int) -> date | None:
    try:
        return date(year, month, day)
    except ValueError:
        return None


def extract_date(text: str | None) -> date | None:
    """Бичвэрээс огноо ялгаж авна.  Олдохгүй бол ``None``."""
    if not text:
        return None
    ymd = DATE_YMD_RE.search(text)
    if ymd:
        found = _make_date(int(ymd.group(1)), int(ymd.group(2)), int(ymd.group(3)))
        if found:
            return found
    # Монголд DD/MM/YYYY хэлбэр түгээмэл.
    dmy = DATE_DMY_RE.search(text)
    if dmy:
        found = _make_date(int(dmy.group(3)), int(dmy.group(2)), int(dmy.group(1)))
        if found:
            return found
    return None


def parse_filename(filename: str | None) -> tuple[str, str]:
    """Файлын нэрнээс (валют, дансны дугаар) ялгана."""
    match = FILENAME_RE.search(filename or "")
    if match:
        return match.group(1).upper(), match.group(2)
    return "MNT", ""


def _to_decimal(value: Any) -> Decimal:
    """Мөнгөн нүдийг ``Decimal`` болгоно.  Танигдахгүй бол 0."""
    if value is None or value == "":
        return ZERO
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return ZERO
    text = re.sub(r"[\s,' ]", "", str(value))
    if not text or text.lower() == "nan":
        return ZERO
    try:
        return Decimal(text)
    except InvalidOperation:
        return ZERO


def _to_datetime(value: Any) -> datetime | None:
    """Огнооны нүдийг ``datetime`` болгоно."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = str(value or "").strip()
    if not text:
        return None
    found = extract_date(text)
    if found:
        return datetime(found.year, found.month, found.day)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _clean_counterpart(value: Any) -> str:
    """Харьцсан дансыг цэвэрлэнэ: 5303363476.0 → "5303363476"."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    # "5303363476.0" маягийн бичвэр
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text[:64]


@dataclass
class _Columns:
    date: int = 0
    debit: int = 3
    credit: int = 4
    desc: int = 6
    counterpart: int = 7
    header_row: int = 1


def _detect_columns(rows: list[list[Any]]) -> _Columns:
    """Гарчгийн мөрийг сканнердаж баганы индексүүдийг олно.

    Дебит/Кредит баганы дараалал солигдсон ч зөв танихыг зорино.  Хоёулаа
    олдсон мөрийг л гарчиг гэж үзнэ.
    """
    cols = _Columns()
    for header_row in range(min(4, len(rows))):
        header = rows[header_row]
        found: dict[str, int] = {}
        for index, value in enumerate(header):
            if not isinstance(value, str):
                continue
            text = value.strip().lower()
            if not text:
                continue
            if "дебит" in text or "debit" in text:
                found["debit"] = index
            elif "кредит" in text or "credit" in text:
                found["credit"] = index
            elif ("огноо" in text and ("гүйлгээ" in text or index <= 1)) or text == "date":
                found.setdefault("date", index)
            elif "утга" in text or "тайлбар" in text or "description" in text:
                found["desc"] = index
            elif ("харьцсан" in text and "данс" in text) or "counterparty" in text:
                found["counterpart"] = index
        if "debit" in found and "credit" in found:
            return _Columns(
                date=found.get("date", cols.date),
                debit=found["debit"],
                credit=found["credit"],
                desc=found.get("desc", cols.desc),
                counterpart=found.get("counterpart", cols.counterpart),
                header_row=header_row,
            )
    return cols


def _statement_period(first_row: list[Any]) -> tuple[date | None, date | None]:
    """Эхний мөрөөс хуулгын интервалыг ялгана.

    Хэвлэсэн огноо ба интервал хоёулаа байдаг тул сүүлийн 1-2 огноог сонгоно.
    """
    parts: list[str] = []
    for value in first_row:
        if value is None:
            continue
        if isinstance(value, datetime):
            parts.append(value.date().isoformat())
        elif isinstance(value, date):
            parts.append(value.isoformat())
        else:
            parts.append(str(value))
    joined = " ".join(parts)
    found = re.findall(r"\d{4}[-/.]\d{2}[-/.]\d{2}", joined)
    if len(found) >= 2:
        return extract_date(found[-2]), extract_date(found[-1])
    if len(found) == 1:
        one = extract_date(found[0])
        return one, one
    return None, None


def parse_statement(content: bytes, filename: str) -> ParsedStatement:
    """Excel хуулгыг задалж гүйлгээний мөрүүдийг буцаана."""
    currency, account_number = parse_filename(filename)
    result = ParsedStatement(
        account_number=account_number, currency=currency, filename=(filename or "")[:255]
    )

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — файл эвдэрсэн, эсвэл Excel биш
        raise ValueError("Excel файлыг уншиж чадсангүй") from exc

    try:
        sheet = workbook.worksheets[0] if workbook.worksheets else None
        if sheet is None:
            return result
        rows: list[list[Any]] = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if not rows:
        return result

    result.date_from, result.date_to = _statement_period(rows[0])
    cols = _detect_columns(rows)

    width = max(cols.date, cols.debit, cols.credit, cols.desc, cols.counterpart) + 1

    def cell(row: list[Any], index: int) -> Any:
        return row[index] if index < len(row) else None

    for row in rows[cols.header_row + 1 :]:
        if len(row) < width:
            row = list(row) + [None] * (width - len(row))
        first = cell(row, cols.date)
        if first is None:
            continue
        # Нийлбэр мөрийг алгасна.
        if isinstance(first, str) and (not first.strip() or first.strip().startswith("Нийт")):
            continue

        txn_date = _to_datetime(first)
        if txn_date is None:
            continue  # огноо танигдахгүй бол гүйлгээний мөр биш

        # Зарим хуулгад дебит сөрөг утгатай ирдэг тул абсолют утгыг авна.
        debit = abs(_to_decimal(cell(row, cols.debit)))
        credit = abs(_to_decimal(cell(row, cols.credit)))

        raw_desc = cell(row, cols.desc)
        desc = "" if raw_desc is None else str(raw_desc).strip()
        if desc.lower() == "nan":
            desc = ""

        result.transactions.append(
            ParsedTransaction(
                txn_date=txn_date,
                debit=debit,
                credit=credit,
                bank_description=desc,
                bank_counterpart=_clean_counterpart(cell(row, cols.counterpart)),
                is_fee=is_fee_description(desc),
            )
        )

    return result
