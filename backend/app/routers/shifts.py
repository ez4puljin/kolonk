"""Ээлжийн API (WP5) — нээх, хаах, хянах, тайлагнах.

Excel тайланг энэ файлд openpyxl-ээр шууд барина (гадны excel_service-ээс хамаарахгүй),
барих ажлыг `run_in_threadpool`-оор гүйцэтгэнэ.
"""

from __future__ import annotations

import io
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any, Optional
from urllib.parse import quote
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status as http_status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.concurrency import run_in_threadpool

from app.config import settings
from app.database import get_db
from app.deps import get_current_user, require_permission
from app.enums import ShiftStatus
from app.models.user import User
from app.models.shift import ShiftAttachment
from app.schemas.shift import (
    ClosingApprovalIn,
    ClosingCorrectIn,
    CurrentShiftOut,
    DailyCloseIn,
    DailyPreviewIn,
    PriceMarkIn,
    PriceMarkOut,
    ShiftAttachmentOut,
    ShiftCloseIn,
    ShiftListOut,
    ShiftOpenIn,
    ShiftReportOut,
)
from app.services import attendant_service, shift_service
from app.services.audit_service import audit

router = APIRouter(prefix="/api", tags=["shifts"])

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SHEET_TITLE = "Ээлжийн тайлан"

MONEY_FMT = "#,##0.00"
LITER_FMT = "#,##0.000"

TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
HEADER_FONT = Font(bold=True)
LABEL_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", start_color="FFE2E8F0", end_color="FFE2E8F0")

try:  # контейнерт tzdata байхгүй байж болно
    STATION_TZ: Any = ZoneInfo(settings.tz)
except Exception:  # noqa: BLE001  # pragma: no cover
    STATION_TZ = UTC


# --------------------------------------------------------------------------- #
# Endpoints
# --------------------------------------------------------------------------- #
@router.get("/shifts/current", response_model=Optional[CurrentShiftOut])  # noqa: UP007
async def current_shift(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any] | None:
    """Нээлттэй ээлж болон түүний шууд хураангуй (байхгүй бол null).

    Салбартай хэрэглэгч зөвхөн ӨӨРИЙН салбарын ээлжийг хардаг — хөрш
    салбарын ээлж дээр санамсаргүй ажиллахаас сэргийлнэ.
    """
    shift = await shift_service.open_shift_for_user(db, user)
    if shift is None:
        return None
    return await shift_service.current_shift_summary(db, shift)


@router.post("/shifts/open", response_model=CurrentShiftOut, status_code=http_status.HTTP_201_CREATED)
async def open_shift(
    payload: ShiftOpenIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("shifts.open")),
) -> dict[str, Any]:
    """Ээлж нээх — эхний бэлэн мөнгө, савны хэмжилт, хошууны тоолуурын заалт."""
    shift = await shift_service.open_shift(
        db,
        user,
        opening_cash=payload.opening_cash,
        tank_dips=payload.tank_dips,
        totalizer_readings=payload.totalizer_readings,
    )
    return await shift_service.current_shift_summary(db, shift)


@router.post("/shifts/{shift_id}/close", response_model=ShiftReportOut)
async def close_shift(
    shift_id: uuid.UUID,
    payload: ShiftCloseIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("shifts.close")),
) -> dict[str, Any]:
    """Ээлж хаах — тооцоо нийлүүлж, зөрүүг журналд бичээд бүрэн тайланг буцаана."""
    shift = await shift_service.get_shift(db, shift_id)
    return await shift_service.close_shift(
        db,
        user,
        shift=shift,
        declared_cash=payload.declared_cash,
        tank_dips=payload.tank_dips,
        totalizer_readings=payload.totalizer_readings,
        note=payload.note,
    )


# --------------------------------------------------------------------------- #
# Түгээгчийн өдрийн ээлж — зураг, үнийн тэмдэглэл, өдрийн хаалт
# --------------------------------------------------------------------------- #
from pathlib import Path  # noqa: E402

from sqlalchemy import select as sa_select  # noqa: E402

ATTACHMENT_DIR = Path("uploads") / "shift_photos"
MAX_PHOTO_BYTES = 10 * 1024 * 1024
#: Зөвшөөрөгдсөн файлын төрлүүд — агуулгын эхний байтаар шалгана.
PHOTO_SIGNATURES: dict[str, bytes] = {
    "image/jpeg": bytes([0xFF, 0xD8, 0xFF]),
    "image/png": bytes([0x89]) + b"PNG",
    "image/webp": b"RIFF",
    "application/pdf": b"%PDF",
}
PHOTO_EXTENSIONS: dict[str, str] = {
    "image/jpeg": "jpg",
    "image/png": "png",
    "image/webp": "webp",
    "application/pdf": "pdf",
}
ATTACHMENT_KINDS = {"open", "close", "settlement", "price_mark"}


def _attachment_out(row: ShiftAttachment) -> dict[str, Any]:
    return {
        "id": row.id,
        "kind": row.kind,
        "ref_id": row.ref_id,
        "original_name": row.original_name,
        "content_type": row.content_type,
        "size_bytes": row.size_bytes,
        "created_at": row.created_at,
    }


@router.post(
    "/shifts/{shift_id}/attachments",
    response_model=ShiftAttachmentOut,
    status_code=http_status.HTTP_201_CREATED,
)
async def upload_shift_attachment(
    shift_id: uuid.UUID,
    kind: str = Query(default="open"),
    ref_id: uuid.UUID | None = Query(default=None),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("shifts.open", "shifts.close")),
) -> dict[str, Any]:
    """Ээлжид зураг хавсаргана (миль, кассын тоолол, settlement-ийн баримт)."""
    if kind not in ATTACHMENT_KINDS:
        raise HTTPException(status_code=422, detail="Хавсралтын төрөл буруу байна")
    shift = await shift_service.get_shift(db, shift_id)

    content = await file.read()
    if not content:
        raise HTTPException(status_code=422, detail="Файл хоосон байна")
    if len(content) > MAX_PHOTO_BYTES:
        raise HTTPException(status_code=422, detail="Файл 10MB-аас хэтэрч байна")
    # Агуулгаар нь шалгана — өргөтгөлд итгэхгүй.
    content_type = next(
        (ctype for ctype, sig in PHOTO_SIGNATURES.items() if content.startswith(sig)), None
    )
    if content_type is None:
        raise HTTPException(
            status_code=422, detail="Зөвхөн JPG/PNG/WEBP зураг эсвэл PDF хавсаргана"
        )

    file_name = f"{uuid.uuid4().hex}.{PHOTO_EXTENSIONS[content_type]}"
    target_dir = ATTACHMENT_DIR / str(shift.id)
    target_dir.mkdir(parents=True, exist_ok=True)
    (target_dir / file_name).write_bytes(content)

    row = ShiftAttachment(
        shift_id=shift.id,
        kind=kind,
        ref_id=ref_id,
        file_name=file_name,
        original_name=(file.filename or "")[:255],
        content_type=content_type,
        size_bytes=len(content),
        uploaded_by=user.id,
    )
    db.add(row)
    await db.flush()
    await audit(
        db,
        user_id=user.id,
        action="shift.attachment",
        entity_type="shift",
        entity_id=shift.id,
        after={"kind": kind, "file": file_name, "size": len(content)},
    )
    return _attachment_out(row)


@router.get("/shifts/{shift_id}/attachments", response_model=list[ShiftAttachmentOut])
async def list_shift_attachments(
    shift_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
) -> list[dict[str, Any]]:
    rows = (
        await db.scalars(
            sa_select(ShiftAttachment)
            .where(ShiftAttachment.shift_id == shift_id)
            .order_by(ShiftAttachment.created_at)
        )
    ).all()
    return [_attachment_out(r) for r in rows]


@router.get("/shifts/{shift_id}/attachments/{attachment_id}/file")
async def download_shift_attachment(
    shift_id: uuid.UUID,
    attachment_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
) -> FileResponse:
    row = await db.get(ShiftAttachment, attachment_id)
    if row is None or row.shift_id != shift_id:
        raise HTTPException(status_code=404, detail="Хавсралт олдсонгүй")
    path = ATTACHMENT_DIR / str(shift_id) / row.file_name
    if not path.exists():
        raise HTTPException(status_code=404, detail="Файл олдсонгүй")
    return FileResponse(
        path, media_type=row.content_type, filename=row.original_name or row.file_name
    )


@router.post("/shifts/{shift_id}/price-marks", response_model=PriceMarkOut, status_code=201)
async def add_price_mark(
    shift_id: uuid.UUID,
    payload: PriceMarkIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("shifts.open", "shifts.close")),
) -> dict[str, Any]:
    """Өдрийн дундуур үнэ өөрчлөгдөхөд шинэ үнэ аль мильд эхэлснийг тэмдэглэнэ."""
    shift = await shift_service.get_shift(db, shift_id)
    mark = await attendant_service.add_price_mark(
        db,
        user,
        shift,
        nozzle_id=payload.nozzle_id,
        reading=payload.reading,
        new_price=payload.new_price,
        note=payload.note,
    )
    rows = await attendant_service.price_marks_out(db, shift)
    return next(r for r in rows if r["id"] == mark.id)


@router.get("/shifts/{shift_id}/price-marks", response_model=list[PriceMarkOut])
async def list_price_marks(
    shift_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(get_current_user),
) -> list[dict[str, Any]]:
    shift = await shift_service.get_shift(db, shift_id)
    return await attendant_service.price_marks_out(db, shift)


@router.post("/shifts/{shift_id}/daily-preview")
async def daily_preview(
    shift_id: uuid.UUID,
    payload: DailyPreviewIn,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_permission("shifts.close")),
) -> dict[str, Any]:
    """Хаалтын өмнөх миль×үнэ тулгалт — юу ч бичихгүй."""
    shift = await shift_service.get_shift(db, shift_id)
    return await attendant_service.daily_preview(db, shift, payload.totalizer_readings)


@router.post("/shifts/{shift_id}/daily-close", response_model=ShiftReportOut)
async def daily_close(
    shift_id: uuid.UUID,
    payload: DailyCloseIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("shifts.close")),
) -> dict[str, Any]:
    """Түгээгчийн өдрийн хаалт — бүх бүртгэл + ээлж хаах нэг дор."""
    shift = await shift_service.get_shift(db, shift_id)
    return await attendant_service.daily_close(db, user, shift, payload)


@router.get("/shifts", response_model=ShiftListOut)
async def list_shifts(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    status: str | None = Query(default=None),
    branch_id: uuid.UUID | None = Query(default=None, description="Зөвхөн энэ салбарын ээлж"),
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("shifts.view_all")),
) -> dict[str, Any]:
    """Ээлжийн жагсаалт — огноо/төлөв/салбараар шүүх, хуудаслах."""
    if status is not None and status not in tuple(ShiftStatus):
        raise HTTPException(status_code=422, detail="Ээлжийн төлөв буруу байна")
    if date_from is not None and date_to is not None and date_from > date_to:
        raise HTTPException(status_code=422, detail="Эхлэх огноо дуусах огнооноос хойш байж болохгүй")
    return await shift_service.list_shifts(
        db,
        date_from=date_from,
        date_to=date_to,
        status=status,
        branch_id=branch_id,
        limit=limit,
        offset=offset,
    )


@router.get("/shifts/daily-closings")
async def daily_closings(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    branch_id: list[uuid.UUID] | None = Query(default=None),
    attendant_id: list[uuid.UUID] | None = Query(default=None),
    status: str | None = Query(default=None),
    only_variance: bool = Query(default=False),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("shifts.view_all", "shifts.close")),
) -> list[dict[str, Any]]:
    """Ээлжийн тайлан — салбар, ажилтан, огноо, батламжийн төлвөөр шүүнэ."""
    if date_from is not None and date_to is not None and date_from > date_to:
        raise HTTPException(status_code=422, detail="Эхлэх огноо дуусах огнооноос хойш байж болохгүй")
    if status is not None and status not in {"approved", "pending"}:
        raise HTTPException(status_code=422, detail="Батламжийн төлөв буруу байна")
    return await attendant_service.daily_closings_list(
        db,
        date_from=date_from,
        date_to=date_to,
        branch_ids=branch_id or None,
        attendant_ids=attendant_id or None,
        status=status,
        only_variance=only_variance,
    )


@router.post("/shifts/{shift_id}/closing/correct")
async def correct_closing(
    shift_id: uuid.UUID,
    payload: ClosingCorrectIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("shifts.approve")),
) -> dict[str, Any]:
    """Тоолсон бэлэн мөнгийг засаж, кассын зөрүүг дахин бичнэ (нягтлан)."""
    return await attendant_service.correct_declared_cash(
        db,
        user,
        shift_id=shift_id,
        declared_cash=payload.declared_cash,
        note=payload.note,
    )


@router.post("/shifts/{shift_id}/closing/approval")
async def set_closing_approval(
    shift_id: uuid.UUID,
    payload: ClosingApprovalIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("shifts.approve")),
) -> dict[str, Any]:
    """Хаалтыг батлах (`approved=true`) эсвэл батламжийг буцаах."""
    return await attendant_service.set_closing_approval(
        db, user, shift_id=shift_id, approved=payload.approved, note=payload.note
    )


@router.get("/shifts/{shift_id}/report.xlsx")
async def shift_report_xlsx(
    shift_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("shifts.view_all", "shifts.close")),
) -> StreamingResponse:
    """Ээлжийн тайланг Excel файлаар татах."""
    shift = await shift_service.get_shift(db, shift_id)
    report = await shift_service.shift_report(db, shift)
    payload = await run_in_threadpool(_build_workbook, report)

    ascii_name = f"shift-{shift.number}.xlsx"
    utf8_name = quote(f"Ээлж-{shift.number}-тайлан.xlsx")
    disposition = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{utf8_name}"
    return StreamingResponse(
        io.BytesIO(payload),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": disposition},
    )


@router.get("/shifts/{shift_id}/report", response_model=ShiftReportOut)
async def shift_report(
    shift_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("shifts.view_all", "shifts.close")),
) -> dict[str, Any]:
    """Ээлжийн бүрэн тайлан."""
    shift = await shift_service.get_shift(db, shift_id)
    report = await shift_service.shift_report(db, shift)
    # Түгээгчийн өдрийн хаалттай бол миль×үнэ тооцоог хамт өгнө.
    report["daily"] = await attendant_service.closing_out(db, shift)
    return report


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def _fmt_dt(value: Any) -> str:
    """Excel timezone-той огноо дэмждэггүй тул текстээр (станцын цагаар) бичнэ."""
    if not isinstance(value, datetime):
        return ""
    moment = value if value.tzinfo is not None else value.replace(tzinfo=UTC)
    return moment.astimezone(STATION_TZ).strftime("%Y-%m-%d %H:%M")


def _num(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _build_workbook(report: dict[str, Any]) -> bytes:
    """Ээлжийн тайлангийн ажлын дэвтрийг барина (threadpool дотор ажиллана)."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_TITLE
    ws.sheet_view.showGridLines = False

    cursor = {"row": 1}

    def cell(col: int, value: Any, *, font: Font | None = None, fmt: str | None = None,
             fill: PatternFill | None = None) -> None:
        target = ws.cell(row=cursor["row"], column=col, value=value)
        if font is not None:
            target.font = font
        if fmt is not None:
            target.number_format = fmt
        if fill is not None:
            target.fill = fill
        target.alignment = Alignment(vertical="center")

    def newline(count: int = 1) -> None:
        cursor["row"] += count

    def title(text: str) -> None:
        cell(1, text, font=TITLE_FONT)
        newline()

    def section(text: str) -> None:
        newline()
        cell(1, text, font=SECTION_FONT)
        newline()

    def kv(label: str, value: Any, fmt: str | None = None) -> None:
        cell(1, label, font=LABEL_FONT)
        cell(2, value, fmt=fmt)
        newline()

    def header(columns: list[str]) -> None:
        for index, name in enumerate(columns, start=1):
            cell(index, name, font=HEADER_FONT, fill=HEADER_FILL)
        newline()

    def data(values: list[Any], formats: list[str | None]) -> None:
        for index, value in enumerate(values, start=1):
            fmt = formats[index - 1] if index - 1 < len(formats) else None
            cell(index, value, fmt=fmt)
        newline()

    meta = report.get("shift", {})
    sales = report.get("sales", {})
    cash = report.get("cash", {})
    profit = report.get("profit", {})

    # --- Толгой ---
    title(f"{settings.station_name} — ЭЭЛЖИЙН ТАЙЛАН")
    kv("Ээлжийн дугаар", meta.get("number"))
    kv("Төлөв", meta.get("status_name"))
    kv("Нээсэн", _fmt_dt(meta.get("opened_at")))
    kv("Нээсэн ажилтан", meta.get("opened_by_name") or "")
    kv("Хаасан", _fmt_dt(meta.get("closed_at")))
    kv("Хаасан ажилтан", meta.get("closed_by_name") or "")
    kv("Тэмдэглэл", meta.get("note") or "")

    # --- Борлуулалт ---
    section("БОРЛУУЛАЛТЫН ДҮН")
    kv("Гүйлгээний тоо", sales.get("count", 0))
    kv("Нийт борлуулалт", _num(sales.get("gross_total")), MONEY_FMT)
    kv("Үүнээс НӨАТ", _num(sales.get("vat_total")), MONEY_FMT)
    kv("НӨАТ-гүй дүн", _num(sales.get("net_total")), MONEY_FMT)
    kv("Түлшний борлуулалт", _num(sales.get("fuel_amount")), MONEY_FMT)
    kv("Түлшний нийт литр", _num(sales.get("fuel_liters")), LITER_FMT)
    kv("Дэлгүүрийн борлуулалт", _num(sales.get("store_amount")), MONEY_FMT)

    # --- Төлбөрийн хэлбэр ---
    section("ТӨЛБӨРИЙН ХЭЛБЭРЭЭР")
    header(["Хэлбэр", "Гүйлгээ", "Дүн"])
    for row in sales.get("by_tender", []):
        data([row.get("method_name"), row.get("count", 0), _num(row.get("amount"))], [None, None, MONEY_FMT])

    # --- Түлшний төрөл ---
    section("ТҮЛШНИЙ ТӨРЛӨӨР")
    header(["Код", "Нэр", "Литр", "Дүн"])
    for row in report.get("fuels", []):
        data(
            [row.get("code"), row.get("name"), _num(row.get("liters")), _num(row.get("amount"))],
            [None, None, LITER_FMT, MONEY_FMT],
        )

    # --- Насос/хошуу ---
    section("НАСОС / ХОШУУНЫ ТООЛУУР")
    header(
        [
            "Насос",
            "Хошуу",
            "Түлш",
            "Нээлтийн заалт",
            "Хаалтын заалт",
            "Тоолуурын зөрүү (л)",
            "Борлуулсан (л)",
            "Дүн",
        ]
    )
    for row in report.get("nozzles", []):
        data(
            [
                f"{row.get('pump_number')} — {row.get('pump_name')}",
                row.get("nozzle_number"),
                row.get("fuel_name"),
                _num(row.get("opening_reading")),
                _num(row.get("closing_reading")),
                _num(row.get("reading_delta_l")),
                _num(row.get("sold_liters")),
                _num(row.get("sold_amount")),
            ],
            [None, None, None, LITER_FMT, LITER_FMT, LITER_FMT, LITER_FMT, MONEY_FMT],
        )

    # --- Сав ---
    section("САВНЫ ХЭМЖИЛТ БА ЗӨРҮҮ")
    header(
        [
            "Сав",
            "Түлш",
            "Нээлтийн хэмжилт",
            "Хаалтын хэмжилт",
            "Дэвтрийн үлдэгдэл",
            "Зөрүү (л)",
            "Зөрүүний өртөг",
        ]
    )
    for row in report.get("tanks", []):
        data(
            [
                row.get("tank_name"),
                row.get("fuel_name"),
                _num(row.get("open_dip")),
                _num(row.get("close_dip")),
                _num(row.get("book_liters")),
                _num(row.get("variance_l")),
                _num(row.get("variance_value")),
            ],
            [None, None, LITER_FMT, LITER_FMT, LITER_FMT, LITER_FMT, MONEY_FMT],
        )

    # --- Касс ---
    section("БЭЛЭН МӨНГӨНИЙ ТООЦОО")
    kv("Эхний үлдэгдэл", _num(cash.get("opening_cash")), MONEY_FMT)
    kv("Бэлэн борлуулалт", _num(cash.get("cash_sales")), MONEY_FMT)
    kv("Бэлнээр буцаасан", _num(cash.get("refunds")), MONEY_FMT)
    kv("Байвал зохих", _num(cash.get("expected_cash")), MONEY_FMT)
    kv("Тоолсон бэлэн", _num(cash.get("declared_cash")), MONEY_FMT)
    kv("Илүүдэл (+) / Дутагдал (−)", _num(cash.get("cash_over_short")), MONEY_FMT)

    # --- Буцаалт ---
    section("БУЦААЛТ")
    header(["Борлуулалт №", "Дүн", "Хэлбэр", "Төлөв", "Шалтгаан"])
    for row in report.get("refunds", []):
        data(
            [
                row.get("sale_number"),
                _num(row.get("amount")),
                row.get("refund_method_name"),
                row.get("status_name"),
                row.get("reason") or "",
            ],
            [None, MONEY_FMT, None, None, None],
        )

    # --- Ашиг ---
    section("ӨРТӨГ БА АШИГ")
    kv("Борлуулалтын орлого (НӨАТ-гүй)", _num(profit.get("revenue_net")), MONEY_FMT)
    kv("Борлуулсан бүтээгдэхүүний өртөг", _num(profit.get("cogs_total")), MONEY_FMT)
    kv("Нийт ашиг", _num(profit.get("gross_profit")), MONEY_FMT)
    kv("Ашгийн хувь (%)", _num(profit.get("margin_pct")), MONEY_FMT)

    # --- Журналын бичилт ---
    entries = report.get("posted_entries", [])
    if entries:
        section("ХИЙГДСЭН ЖУРНАЛЫН БИЧИЛТ")
        header(["Бичилт №", "Үйл явдал", "Тайлбар", "Дүн"])
        for row in entries:
            data(
                [row.get("entry_no"), row.get("event_type"), row.get("description"), _num(row.get("amount"))],
                [None, None, None, MONEY_FMT],
            )

    for column, width in (("A", 34), ("B", 22), ("C", 18), ("D", 18), ("E", 18), ("F", 20), ("G", 20), ("H", 18)):
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
